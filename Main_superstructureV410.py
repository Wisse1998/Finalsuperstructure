# -*- coding: utf-8 -*-
"""
Created on Wed Jun  1 12:37:15 2022

@author: Gebruiker
"""

import pandas as pd
from Superstructure_classV410 import Superstructure
from Superstructure_modelV410 import Superstructure_model



# 'C:/Users/Gebruiker/iCloudDrive/Master/Thesis/Mathematical formulation/SSdataV1.xlsx'
# 'C:/Users/Gebruiker/iCloudDrive/Master/Thesis/Python/Vincent\SS_dataVINCENT.xlsx'

xlsx_file = 'C:/Users/Gebruiker/iCloudDrive/Master/Thesis/Mathematical formulation/SSdataV406.xlsx'
output_file = 'C:/Users/Gebruiker/iCloudDrive/Master/Thesis/Python/Scenario4/ResultsS4.xlsx'


Superstructure = Superstructure(xlsx_file)
Superstructure.get_SF(xlsx_file)
Superstructure.get_EC(xlsx_file)
Superstructure.get_F0(xlsx_file)
Superstructure.get_RatiomR(xlsx_file)
Superstructure.get_CF(xlsx_file)
Superstructure.get_SEL(xlsx_file)
Superstructure.get_MW(xlsx_file)
Superstructure.get_ALPHA(xlsx_file)
Superstructure.get_UB(xlsx_file)
Superstructure.get_LB(xlsx_file)
Superstructure.get_Slope(xlsx_file)
Superstructure.get_constant(xlsx_file)
Superstructure.get_Uprice(xlsx_file)
Superstructure.get_Tau(xlsx_file)
Superstructure.get_Distance(xlsx_file)
Superstructure.get_SupplyLoad(xlsx_file)
Superstructure.get_CCost(xlsx_file)
Superstructure.get_Demand(xlsx_file)
Superstructure.get_Marketdis(xlsx_file)



##############################################################
model = Superstructure_model(Superstructure)

Superstructure.get_results(model)


from openpyxl import Workbook
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


filename = output_file

workbook = Workbook()
sheet = workbook.active

#Make numpy arrays for even spacing in excel sheet to write equipment data
start_col = 1
step_col = 5
col = np.arange(start_col, len(Superstructure.a)*step_col+start_col, step_col)
start_row = 5
step_row = 8
row = np.arange(start_row, len(Superstructure.k)*step_row+start_row, step_row)

h_spacer = 10

#Size of the data square for equpiment types that is colored
h_size = 3
v_size = 8

cell = get_column_letter(start_col) + str(start_row + h_spacer)
cell1 = get_column_letter(start_col+5) + str(start_row + h_spacer)

sheet[get_column_letter(start_col) + str(start_row + h_spacer + 2)] = 'AIC'
sheet[get_column_letter(start_col+1) + str(start_row + h_spacer + 2)] = Superstructure.AIC / 1000

workbook.save(filename = filename)



